from datetime import date
from openpyxl import Workbook

class Vendedor:
  # atributos da classe
  def __init__(self, nome, email, data_venda, valor_venda):
    self.nome = nome
    self.email = email
    self.data_venda = data_venda
    self.valor_venda = valor_venda
    self.comissao = 0

  def calcular_comissao(self):
    nomeVendedorDoEmail = self.email.split('@')[0]

    # comissao dos vendedores
    vendedoresAtivos = {
      'marconnirodrigues': 1,
      'katianeves': 40,
      'thiagogiardini': 40,
      'wagneroliveira': 30,
    }
    
    # pega o vendedor da lista, se não tiver presente ele recebe 0% de comissão
    percentualDeComissao = vendedoresAtivos.get(nomeVendedorDoEmail, 0)
    
    self.comissao = (percentualDeComissao / 100) * self.valor_venda

  # to string
  def valor_comissao(self):
    return f'O vendedor {self.nome} vai receber R${self.comissao:.2f} de comissão.'

def criarPlanilha(vendedores, nomeDoArquivo="Vendedores.xlsx"):
  workbook = Workbook()  # Correção aqui
  
  # Preencher a planilha para cada vendedor
  for vendedor in vendedores:
    # Criar uma nova planilha para cada vendedor
    sheet = workbook.create_sheet(title=f"{vendedor.nome} - Semana {vendedor.data_venda.isocalendar()[1]}")
    
    # Adicionar cabeçalhos
    sheet.append(["Nome do Vendedor", "Total Vendido", "Comissão"])
    
    # Calcular a comissão
    vendedor.calcular_comissao()
    
    # Preencher a planilha com dados
    sheet.append([vendedor.nome, vendedor.valor_venda, vendedor.comissao])

  # Remover a planilha padrão criada automaticamente
  if "Sheet" in workbook.sheetnames:
    del workbook["Sheet"]

  # Salvar a planilha
  workbook.save(nomeDoArquivo)

vendedores = [
  Vendedor("Marconni Rodrigues", "marconnirodrigues@example.com", date.today(), 1000),
  Vendedor("Katiane Neves", "katianeves@example.com", date.today(), 2000),
  Vendedor("Thiago Giardini", "thiagogiardini@example.com", date.today(), 1500),
  Vendedor("Wagner Oliveira", "wagneroliveira@example.com", date.today(), 2500)
]

# Criar a planilha
criarPlanilha(vendedores)

print("Planilha criada com sucesso!")
