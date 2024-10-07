
from planilha_module import Planilha
from openpyxl import Workbook, load_workbook
import os

class Vendedor:
    def __init__( self, vendedor_nome, vendedor_email, venda_data, venda_valor_total ):
        self.vendedor_nome     = vendedor_nome 
        self.vendedor_email    = vendedor_email
        self.venda_data        = venda_data
        self.venda_valor_total = venda_valor_total
        self.vendedor_comissao = 0.0
    # __init__ ( )

    """ Vendedor toString """
    def __str__( self ) -> str:
        return (f"Vendedor:\n"
                f"Nome................: {self.vendedor_nome}\n"
                f"E-mail..............: {self.vendedor_email}\n"
                f"Data da Venda.......: {self.venda_data}\n"
                f"Valor Total da Venda: {self.venda_valor_total}\n"
                f"Valor da Comissão...: {self.vendedor_comissao}\n"
                )
    # __str__ ( )
    
    """ Calcular a comissão com base no vendedor """
    def calcular_comissao( self ):
        nomeVendedorDoEmail = self.vendedor_email.split('@')[0]

        # Comissão dos vendedores
        vendedoresAtivos = {
            'marconnirodrigues': 100,
            'katianeves': 40,
            'thiagogiardini': 40,
            'wagneroliveira': 30,
        }

        percentualDeComissao = vendedoresAtivos.get(nomeVendedorDoEmail, 0)
        self.vendedor_comissao = (percentualDeComissao / 100) * self.venda_valor_total
    # end calcular_comissao ( )

    """ Atualizar a planilha de vendedores, adicionando os valores semanalmente. """
    def gravar_planilha_vendedores( vendedores ):
        output_filepath = os.path.join("data", "processed", "Vendedores.xlsx")
        
        try:
            try:
                workbook = load_workbook(output_filepath)
            except FileNotFoundError:
                workbook = Workbook()

            for res in vendedores:                
                semana_venda = res.venda_data.isocalendar()[1]
                nomes = res.vendedor_nome.split()

                nome_planilha = f"{nomes[0]}"
                if len(nomes) > 1:
                    nome_planilha += f"{nomes[1]}"
                nome_planilha += f" - Semana {semana_venda}"

                # Criar nova planilha se não existir
                if nome_planilha not in workbook.sheetnames:
                    sheet = workbook.create_sheet(title=nome_planilha)
                    sheet.append(["NOME", "EMAIL", "DATA DA VENDA", "VALOR TOTAL DA VENDA", "COMISSÃO"])
                else:
                    sheet = workbook[nome_planilha]

                # Verificar se há mudança de semana e adicionar linha de "X" se necessário
                if sheet.max_row > 1:
                    ultima_semana = sheet.cell(row=sheet.max_row, column=1).value
                if ultima_semana != semana_venda:
                    sheet.append(["X"] * 5)
                
                res.calcular_comissao( )
                
                # Preencher a planilha com os dados
                sheet.append([
                    res.vendedor_nome,
                    res.vendedor_email,
                    res.venda_data,
                    res.venda_valor_total,
                    res.vendedor_comissao
                ])

            # Remover planilha padrão "Sheet" se ela existir
            if "Sheet" in workbook.sheetnames:
                del workbook["Sheet"]

            workbook.save(output_filepath)
            print(f"Planilha de vendedores atualizada com sucesso!")
        except Exception as e:
            print(f"Erro ao atualizar a planilha de vendedores: {e}")
    # end gravar_planilha_vendedor ( )
# Vendedor